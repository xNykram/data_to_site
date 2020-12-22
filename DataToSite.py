from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
import os
from openpyxl import load_workbook

class DataToSite:

    def __init__(self, login, password, targetUrl, path):

        self.login = login
        self.password = password
        self.mainsite = targetUrl
        self.excelDocument = None
        self.fieldsCount = 50
        self.driver = webdriver.Chrome(executable_path=path)

    def loginToSite(self):
        # login to hem site
        print("Logging in...")
        self.driver.get(self.mainsite)
        loginInput = self.driver.find_element_by_id("input_1")
        loginInput.send_keys(self.login)
        passwordInput = self.driver.find_element_by_name("password")
        passwordInput.send_keys(self.password)
        loginButton = self.driver.find_element_by_class_name("credentials_input_submit")
        loginButton.send_keys(Keys.ENTER)
        sleep(5)

    def autoFillInputs(self, amountOfFields, prize, description, listTitle):
        self.fieldsCount = amountOfFields
        companyInput = self.driver.find_element_by_id("idf_invcompany")
        companyInput.send_keys("...")
        sleep(0.1)
        # add rows and fill empty fields
        print("Adding rows...")
        addRow = self.driver.find_element_by_class_name("plus")
        for i in range(self.fieldsCount):
            sleep(0.05)
            addRow.click()
        print("Selecting the correct options...")
        sleep(10)
        for i in range(self.fieldsCount):
                list = self.driver.find_element_by_id("idf_costcenterchoice[{}]".format(i))
                list.send_keys(listTitle[i])
                sleep(0.2)
                list.send_keys(Keys.ARROW_DOWN)
                sleep(0.2)
                list.send_keys(Keys.ENTER)
        print("Entering data to inputs...")
        for i in range(self.fieldsCount):
            insertDataDescription = self.driver.find_element_by_name("f_comment[{}]".format(i))
            insertDataDescription.send_keys(description[i])
        sleep(2)
        print("Adding prize...")
        for i in range(self.fieldsCount):
            insertPrize = self.driver.find_element_by_id("idf_price[{}]".format(i))
            insertPrize.send_keys(prize[i])
        sleep(2)
        print("Adding amount...")
        for index in range(self.fieldsCount):
            insertAmount = self.driver.find_element_by_id("idf_amount[{}]".format(index))
            insertAmount.send_keys("1")
        sleep(99999)

class Excel:
    def __init__(self):
        self.path = None
        self.listTitle = []
        self.description = []
        self.prize = []
        self.amountOfFields = 0
        self.webDriver = None

    def searchExcelFile(self):
        path = os.getcwd()
        path = os.path.join(path, "excel")
        excelFile = str(os.listdir(path)).replace("[","").replace("]","").replace("'","")
        path = os.path.join(path, excelFile)
        self.path = path
    def selectWithPrize(self):
        print("Please enter amount of lines:")
        lines = input()
        print("Please enter location of colum numbers(split , each number):")
        location = input()
        location = eval('[' + location + ']')
        wb = load_workbook(self.path, data_only=True)
        sheet = wb.active
        for j in range(int(lines)):
            for i in range(5, 104):
                cell = sheet.cell(row=i, column=location[j])
                if(cell.value != 0):
                    cellDescription = sheet.cell(row=i, column=location[j]-2)
                    cellCostCenter = sheet.cell(row=i, column=location[j]-1)
                    self.description.append(cellDescription.value)
                    self.listTitle.append(cellCostCenter.value)
                    cellPrize = str(round(cell.value, 2))
                    self.prize.append(cellPrize.replace(".", ","))
                else:
                    continue
        print(self.prize)
        print(self.description)
        print(self.listTitle)
    def setDataToSite(self):
        self.amountOfFields = len(self.prize)
        print(self.amountOfFields)
        DtS = DataToSite("username", "password",
                         "targetlink", self.webDriver)
        DtS.loginToSite()
        DtS.autoFillInputs(self.amountOfFields, self.prize, self.description, self.listTitle)

    def searchForWebDriver(self):
        path = os.getcwd()
        path = os.path.join(path, "webdriver")
        webDriver = str(os.listdir(path)).replace("[", "").replace("]", "").replace("'", "")
        path = os.path.join(path, webDriver)
        self.webDriver = path

if __name__ == '__main__':
    Ex = Excel()
    Ex.searchExcelFile()
    Ex.searchForWebDriver()
    Ex.selectWithPrize()
    Ex.setDataToSite()